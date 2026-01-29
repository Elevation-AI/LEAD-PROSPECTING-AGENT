"""
Flask UI Application for Agent01 Lead Prospecting
==================================================
Interactive UI that runs the full Agent01 pipeline with user input
UPDATED: Now supports multiple input types (Website, PDF, Raw Text)
"""

from flask import Flask, render_template, request, jsonify, session, send_file
import json
import os
import sys
from datetime import datetime
from pathlib import Path
import secrets
import tempfile
from io import BytesIO

# Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will be disabled.")

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.scraper.website_scraper import WebsiteScraper
from src.icp.icp_generator import ICPGenerator
from src.search.company_finder import ProspectFinder
from src.enrichment.apollo_enricher import ApolloEnricher
from src.utils.helpers import validate_url

# NEW: Import input layer components
from src.input.pdf_extractor import PDFExtractor
from src.input.raw_text_handler import RawTextHandler
from src.input.content_aggregator import ContentAggregator

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Global storage for current session data
session_data = {}


@app.route('/')
def index():
    """Main page - start the pipeline"""
    return render_template('index.html')


@app.route('/api/scrape', methods=['POST'])
def api_scrape():
    """Step 1: Scrape website (legacy - kept for backward compatibility)"""
    try:
        data = request.json
        url = data.get('url', '').strip()

        if not validate_url(url):
            return jsonify({"error": "Invalid URL format. Example: https://asana.com"}), 400

        scraper = WebsiteScraper()
        scraped = scraper.scrape_website(url)

        if not scraped or len(scraped["combined_text"]) < 200:
            return jsonify({"error": "Could not extract useful content from website."}), 400

        # Store in session
        session_id = secrets.token_hex(8)
        session_data[session_id] = {
            'url': url,
            'scraped': scraped,
            'step': 1
        }

        return jsonify({
            "success": True,
            "session_id": session_id,
            "content_length": scraped['content_length'],
            "message": f"Successfully scraped {scraped['content_length']} characters"
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# =============================================================================
# NEW: MULTI-INPUT API ENDPOINTS
# =============================================================================

@app.route('/api/process-inputs', methods=['POST'])
def api_process_inputs():
    """
    NEW: Process multiple input types (Website URL, PDF, Raw Text)
    Combines all provided inputs into a single context for ICP generation.
    """
    try:
        contexts = []
        input_sources = []
        url = ""

        # 1. Handle Website URL (from form data)
        url = request.form.get('url', '').strip()
        if url:
            if not validate_url(url):
                return jsonify({"error": "Invalid URL format. Example: https://asana.com"}), 400

            scraper = WebsiteScraper()
            scraped = scraper.scrape_website(url)

            if scraped and len(scraped.get("combined_text", "")) > 200:
                contexts.append({
                    "source": "website",
                    "content": scraped["combined_text"]
                })
                input_sources.append(f"Website: {url}")

        # 2. Handle PDF Upload
        if 'pdf_file' in request.files:
            pdf_file = request.files['pdf_file']
            if pdf_file and pdf_file.filename:
                # Save temporarily
                temp_dir = tempfile.mkdtemp()
                temp_path = os.path.join(temp_dir, pdf_file.filename)
                pdf_file.save(temp_path)

                try:
                    pdf_extractor = PDFExtractor()
                    pdf_context = pdf_extractor.extract_text(temp_path)
                    if pdf_context and len(pdf_context.get("content", "")) > 50:
                        contexts.append(pdf_context)
                        input_sources.append(f"PDF: {pdf_file.filename}")
                except Exception as e:
                    print(f"PDF extraction error: {e}")
                finally:
                    # Cleanup temp file
                    if os.path.exists(temp_path):
                        os.remove(temp_path)
                    if os.path.exists(temp_dir):
                        os.rmdir(temp_dir)

        # 3. Handle Raw Text
        raw_text = request.form.get('raw_text', '').strip()
        if raw_text and len(raw_text) >= 50:
            try:
                raw_handler = RawTextHandler()
                raw_context = raw_handler.process(raw_text)
                contexts.append(raw_context)
                input_sources.append("Raw Text")
            except Exception as e:
                print(f"Raw text processing error: {e}")

        # Validate we have at least one input
        if not contexts:
            return jsonify({
                "error": "No valid input provided. Please enter a website URL, upload a PDF, or paste raw text (min 50 characters)."
            }), 400

        # Aggregate all contexts
        aggregator = ContentAggregator()
        combined_content = aggregator.aggregate(contexts)

        # Create session
        session_id = secrets.token_hex(8)
        session_data[session_id] = {
            'url': url if url else "multi-input",
            'combined_content': combined_content,
            'input_sources': input_sources,
            'step': 1
        }

        return jsonify({
            "success": True,
            "session_id": session_id,
            "content_length": len(combined_content),
            "input_sources": input_sources,
            "message": f"Successfully processed {len(input_sources)} input source(s): {', '.join(input_sources)}"
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/download/<session_id>', methods=['GET'])
def api_download(session_id):
    """
    Download results as JSON file (legacy endpoint)
    """
    try:
        if session_id not in session_data:
            return jsonify({"error": "Invalid session"}), 400

        sess = session_data[session_id]

        # Build download data
        url = sess.get('url', 'multi-input')
        company_slug = "custom_input"
        if url and url != "multi-input":
            company_slug = url.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0].split(".")[0]

        download_data = {
            "generated_at": datetime.now().isoformat(),
            "input_sources": sess.get('input_sources', [url] if url else []),
            "source_url": url,
            "icp": sess.get('icp', {}),
            "prospects_found": len(sess.get('prospects', [])),
            "prospects": sess.get('prospects', []),
            "enriched_contacts": sess.get('enriched', []),
            "total_contacts": sum(len(c.get("contacts", [])) for c in sess.get('enriched', []))
        }

        # Create temp file for download
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{company_slug}_leads_{timestamp}.json"

        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, filename)

        with open(temp_path, 'w', encoding='utf-8') as f:
            json.dump(download_data, f, indent=4, ensure_ascii=False)

        return send_file(
            temp_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/json'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# =============================================================================
# EXCEL DOWNLOAD ENDPOINTS
# =============================================================================

def style_excel_header(ws, row=1):
    """Apply styling to Excel header row"""
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def auto_adjust_column_width(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width


@app.route('/api/download-prospects-excel/<session_id>', methods=['GET'])
def api_download_prospects_excel(session_id):
    """
    Download prospect companies as Excel file
    """
    if not EXCEL_AVAILABLE:
        return jsonify({"error": "Excel export not available. Please install openpyxl."}), 500

    try:
        if session_id not in session_data:
            return jsonify({"error": "Invalid session"}), 400

        sess = session_data[session_id]
        prospects = sess.get('prospects', [])

        if not prospects:
            return jsonify({"error": "No prospects to download"}), 400

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Prospect Companies"

        # Headers
        headers = ["#", "Company Name", "Domain", "Website", "Why Good Fit", "Confidence"]
        ws.append(headers)
        style_excel_header(ws)

        # Data rows
        for idx, prospect in enumerate(prospects, 1):
            domain = prospect.get('domain', '')
            ws.append([
                idx,
                prospect.get('name', ''),
                domain,
                f"https://{domain}" if domain else '',
                prospect.get('why_good_fit', ''),
                prospect.get('confidence', '')
            ])

        # Auto-adjust columns
        auto_adjust_column_width(ws)

        # Generate filename
        url = sess.get('url', 'multi-input')
        company_slug = "custom_input"
        if url and url != "multi-input":
            company_slug = url.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0].split(".")[0]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{company_slug}_prospects_{timestamp}.xlsx"

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/download-contacts-excel/<session_id>', methods=['GET'])
def api_download_contacts_excel(session_id):
    """
    Download enriched contacts as Excel file
    """
    if not EXCEL_AVAILABLE:
        return jsonify({"error": "Excel export not available. Please install openpyxl."}), 500

    try:
        if session_id not in session_data:
            return jsonify({"error": "Invalid session"}), 400

        sess = session_data[session_id]
        enriched = sess.get('enriched', [])

        if not enriched:
            return jsonify({"error": "No enriched contacts to download"}), 400

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Enriched Contacts"

        # Headers
        headers = [
            "#", "Company Name", "Company Domain", "Contact Name", "Job Title",
            "Email", "LinkedIn URL", "Location", "Headline"
        ]
        ws.append(headers)
        style_excel_header(ws)

        # Data rows
        row_num = 1
        for company in enriched:
            company_name = company.get('company_name', '')
            company_domain = company.get('domain', '')

            for contact in company.get('contacts', []):
                ws.append([
                    row_num,
                    company_name,
                    company_domain,
                    contact.get('name', ''),
                    contact.get('title', ''),
                    contact.get('email', ''),
                    contact.get('linkedin_url', ''),
                    contact.get('location', ''),
                    contact.get('headline', '')
                ])
                row_num += 1

        # Auto-adjust columns
        auto_adjust_column_width(ws)

        # Generate filename
        url = sess.get('url', 'multi-input')
        company_slug = "custom_input"
        if url and url != "multi-input":
            company_slug = url.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0].split(".")[0]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{company_slug}_contacts_{timestamp}.xlsx"

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/download-all-excel/<session_id>', methods=['GET'])
def api_download_all_excel(session_id):
    """
    Download complete results as Excel file with multiple sheets
    """
    if not EXCEL_AVAILABLE:
        return jsonify({"error": "Excel export not available. Please install openpyxl."}), 500

    try:
        if session_id not in session_data:
            return jsonify({"error": "Invalid session"}), 400

        sess = session_data[session_id]

        # Create workbook
        wb = Workbook()

        # ========== Sheet 1: ICP Summary ==========
        ws_icp = wb.active
        ws_icp.title = "ICP Summary"

        icp = sess.get('icp', {})
        icp_rows = [
            ["Field", "Value"],
            ["Seller Business Type", icp.get('seller_business_type', '')],
            ["What They Sell", icp.get('what_they_sell', '')],
            ["Customer Industry", icp.get('customer_industry', '')],
            ["Customer Company Size", icp.get('customer_company_size', '')],
            ["Customer Geography", icp.get('customer_geography', '')],
            ["Target Buyers", ", ".join(icp.get('target_buyers', []))],
            ["Ideal Characteristics", ", ".join(icp.get('ideal_customer_characteristics', []))],
            ["Companies to Avoid", ", ".join(icp.get('avoid_company_types', []))],
        ]

        # Geography details
        geo = icp.get('serviceable_geography', {})
        if geo:
            icp_rows.append(["Geographic Scope", geo.get('scope', '')])
            icp_rows.append(["Countries", ", ".join(geo.get('countries', []))])
            icp_rows.append(["Regions/States", ", ".join(geo.get('states_or_regions', []))])

        for row in icp_rows:
            ws_icp.append(row)

        style_excel_header(ws_icp)
        auto_adjust_column_width(ws_icp)

        # ========== Sheet 2: Prospect Companies ==========
        ws_prospects = wb.create_sheet("Prospect Companies")
        prospects = sess.get('prospects', [])

        prospect_headers = ["#", "Company Name", "Domain", "Website", "Why Good Fit", "Confidence"]
        ws_prospects.append(prospect_headers)
        style_excel_header(ws_prospects)

        for idx, prospect in enumerate(prospects, 1):
            domain = prospect.get('domain', '')
            ws_prospects.append([
                idx,
                prospect.get('name', ''),
                domain,
                f"https://{domain}" if domain else '',
                prospect.get('why_good_fit', ''),
                prospect.get('confidence', '')
            ])

        auto_adjust_column_width(ws_prospects)

        # ========== Sheet 3: Enriched Contacts ==========
        ws_contacts = wb.create_sheet("Enriched Contacts")
        enriched = sess.get('enriched', [])

        contact_headers = [
            "#", "Company Name", "Company Domain", "Contact Name", "Job Title",
            "Email", "LinkedIn URL", "Location", "Headline"
        ]
        ws_contacts.append(contact_headers)
        style_excel_header(ws_contacts)

        row_num = 1
        for company in enriched:
            company_name = company.get('company_name', '')
            company_domain = company.get('domain', '')

            for contact in company.get('contacts', []):
                ws_contacts.append([
                    row_num,
                    company_name,
                    company_domain,
                    contact.get('name', ''),
                    contact.get('title', ''),
                    contact.get('email', ''),
                    contact.get('linkedin_url', ''),
                    contact.get('location', ''),
                    contact.get('headline', '')
                ])
                row_num += 1

        auto_adjust_column_width(ws_contacts)

        # Generate filename
        url = sess.get('url', 'multi-input')
        company_slug = "custom_input"
        if url and url != "multi-input":
            company_slug = url.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0].split(".")[0]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{company_slug}_complete_report_{timestamp}.xlsx"

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/generate-icp', methods=['POST'])
def api_generate_icp():
    """Step 2: Generate ICP - supports both legacy (scraped) and new (combined_content) flows"""
    try:
        data = request.json
        session_id = data.get('session_id')

        if session_id not in session_data:
            return jsonify({"error": "Invalid session. Please start over."}), 400

        sess = session_data[session_id]

        # Support both new multi-input flow and legacy single-URL flow
        if 'combined_content' in sess:
            # New multi-input flow
            content = sess['combined_content']
        elif 'scraped' in sess:
            # Legacy single-URL flow
            content = sess['scraped']["combined_text"]
        else:
            return jsonify({"error": "No content available. Please provide input first."}), 400

        icp_gen = ICPGenerator()
        icp = icp_gen.generate_icp(content)

        sess['icp_original'] = icp
        sess['icp'] = icp
        sess['step'] = 2

        return jsonify({
            "success": True,
            "icp": icp,
            "input_sources": sess.get('input_sources', [])
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/update-icp', methods=['POST'])
def api_update_icp():
    """Step 3: Update ICP with user customizations"""
    try:
        data = request.json
        session_id = data.get('session_id')
        updates = data.get('updates', {})

        if session_id not in session_data:
            return jsonify({"error": "Invalid session"}), 400

        sess = session_data[session_id]
        icp = sess['icp']

        # Apply updates
        geography_changed = False

        if 'countries' in updates and updates['countries']:
            icp['serviceable_geography']['countries'] = updates['countries']
            geography_changed = True

        if 'regions' in updates and updates['regions']:
            icp['serviceable_geography']['states_or_regions'] = updates['regions']
            geography_changed = True

        # Only change scope to custom if geography was actually modified
        if geography_changed:
            icp['serviceable_geography']['scope'] = 'custom'
            # Update geography notes
            notes_parts = []
            if icp['serviceable_geography'].get('countries'):
                notes_parts.append(f"Countries: {', '.join(icp['serviceable_geography']['countries'])}")
            if icp['serviceable_geography'].get('states_or_regions'):
                notes_parts.append(f"Regions: {', '.join(icp['serviceable_geography']['states_or_regions'])}")
            if notes_parts:
                icp['serviceable_geography']['notes'] = " | ".join(notes_parts)

        if 'customer_industry' in updates:
            icp['customer_industry'] = updates['customer_industry']

        if 'target_buyers' in updates:
            icp['target_buyers'] = updates['target_buyers']

        # v2.0: Handle seller_business_type
        if 'seller_business_type' in updates and updates['seller_business_type']:
            icp['seller_business_type'] = updates['seller_business_type']

        # v2.0: Handle avoid_company_types
        if 'avoid_company_types' in updates:
            icp['avoid_company_types'] = updates['avoid_company_types']

        sess['icp'] = icp
        sess['step'] = 3

        return jsonify({
            "success": True,
            "icp": icp
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/find-prospects', methods=['POST'])
def api_find_prospects():
    """Step 4: Find prospect companies"""
    try:
        data = request.json
        session_id = data.get('session_id')

        if session_id not in session_data:
            return jsonify({"error": "Invalid session"}), 400

        sess = session_data[session_id]
        icp = sess['icp']

        finder = ProspectFinder()
        prospects = finder.find_prospects(icp)

        if not prospects:
            return jsonify({"error": "No prospect companies found."}), 400

        sess['prospects'] = prospects
        sess['step'] = 4

        # Filter for display (remove source and confidence)
        prospects_display = [
            {
                "name": p.get("name", ""),
                "domain": p.get("domain", ""),
                "why_good_fit": p.get("why_good_fit", "")
            }
            for p in prospects
        ]

        return jsonify({
            "success": True,
            "prospects_found": len(prospects),
            "prospects": prospects_display
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/enrich', methods=['POST'])
def api_enrich():
    """Step 5: Enrich contacts"""
    try:
        data = request.json
        session_id = data.get('session_id')
        unlock_emails = data.get('unlock_emails', False)
        single_company = data.get('single_company')

        if session_id not in session_data:
            return jsonify({"error": "Invalid session"}), 400

        sess = session_data[session_id]
        icp = sess['icp']

        # If single company enrichment
        if single_company:
            prospects = [single_company]
        else:
            prospects = sess['prospects']

        enricher = ApolloEnricher(unlock_emails=unlock_emails)
        enriched = enricher.enrich(prospects, icp)

        # Store or append enriched data
        if single_company:
            # Append to existing enriched data
            if 'enriched' not in sess:
                sess['enriched'] = []
            sess['enriched'].extend(enriched)
        else:
            # Full enrichment
            sess['enriched'] = enriched

        sess['step'] = 5

        # Filter for display
        enriched_display = [
            {
                "company_name": e.get("company_name", ""),
                "domain": e.get("domain", ""),
                "contacts": [
                    {
                        "name": c.get("name", ""),
                        "title": c.get("title", ""),
                        "email": c.get("email", ""),
                        "linkedin_url": c.get("linkedin_url", ""),
                        "location": c.get("location", ""),
                        "headline": c.get("headline", "")
                    }
                    for c in e.get("contacts", [])
                ]
            }
            for e in enriched
        ]

        total_contacts = sum(len(c.get("contacts", [])) for c in enriched)

        # Only save to file if not single company (save at the end)
        if not single_company:
            save_output(sess)

        return jsonify({
            "success": True,
            "enriched_contacts": enriched_display,
            "total_contacts": total_contacts
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def save_output(sess):
    """Save complete output to JSON file - supports multi-input flow"""
    try:
        output_dir = PROJECT_ROOT / "output"
        output_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        url = sess.get('url', 'multi-input')

        # Handle company slug for multi-input
        if url and url != "multi-input":
            company_slug = url.replace("https://", "").replace("http://", "").replace("www.", "").split("/")[0].split(".")[0]
        else:
            company_slug = "custom_input"

        # Filter ICP for saving (v2.0: includes seller_business_type and avoid_company_types)
        icp_filtered = {
            "seller_business_type": sess['icp'].get("seller_business_type", "unknown"),
            "what_they_sell": sess['icp'].get("what_they_sell", ""),
            "customer_industry": sess['icp'].get("customer_industry", ""),
            "customer_company_size": sess['icp'].get("customer_company_size", ""),
            "target_buyers": sess['icp'].get("target_buyers", []),
            "ideal_customer_characteristics": sess['icp'].get("ideal_customer_characteristics", []),
            "customer_geography": sess['icp'].get("customer_geography", ""),
            "serviceable_geography": sess['icp'].get("serviceable_geography", {}),
            "avoid_company_types": sess['icp'].get("avoid_company_types", []),
        }

        full_output = {
            "generated_at": datetime.now().isoformat(),
            "input_sources": sess.get('input_sources', [url] if url else []),
            "source_url": url,
            "icp": icp_filtered,
            "prospects_found": len(sess.get('prospects', [])),
            "prospects": sess.get('prospects', []),
            "enriched_contacts": sess.get('enriched', []),
            "total_contacts": sum(len(c.get("contacts", [])) for c in sess.get('enriched', []))
        }

        filename = f"{company_slug}_output_{timestamp}.json"
        filepath = output_dir / filename

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(full_output, f, indent=4, ensure_ascii=False)

        print(f"\n Output saved to: {filepath}")
        return str(filepath)

    except Exception as e:
        print(f"Error saving output: {e}")
        return None


@app.route('/api/save-final', methods=['POST'])
def api_save_final():
    """Save final output to file"""
    try:
        data = request.json
        session_id = data.get('session_id')

        if session_id not in session_data:
            return jsonify({"error": "Invalid session"}), 400

        sess = session_data[session_id]
        save_output(sess)

        return jsonify({"success": True, "message": "Output saved successfully"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/health')
def health():
    """Health check"""
    return jsonify({"status": "ok", "timestamp": datetime.now().isoformat()})


if __name__ == '__main__':
    print("\n" + "="*60)
    print("  🎯 AGENT01 INTERACTIVE UI")
    print("="*60)
    print("\n✨ Features:")
    print("  • Real-time pipeline execution")
    print("  • ICP customization interface")
    print("  • Step-by-step workflow")
    print("  • Beautiful results display")
    print("\n🚀 Starting server at http://localhost:5000")
    print("   Press Ctrl+C to stop\n")
    print("="*60 + "\n")

    app.run(debug=True, host='0.0.0.0', port=5000)
